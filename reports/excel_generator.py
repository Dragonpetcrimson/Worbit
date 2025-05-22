"""
reports/excel_generator.py - Excel report generation
"""

import os
import logging
import pandas as pd
import re
import sys
from typing import List, Dict, Any, Optional
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Add parent directory to path to fix module imports
parent_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
if parent_dir not in sys.path:
    sys.path.insert(0, parent_dir)

from reports.base import ReportGenerator, ReportConfig, ReportData, ensure_datetime, sanitize_text

# Try to import path utilities
try:
    from utils.path_utils import (
        get_output_path,
        OutputType,
        normalize_test_id,
        get_standardized_filename
    )
    PATH_UTILS_AVAILABLE = True
except ImportError:
    PATH_UTILS_AVAILABLE = False
    # Define fallback functions
    def normalize_test_id(test_id: str) -> str:
        """Fallback normalize_test_id function."""
        if not test_id:
            return ""
        return test_id if test_id.upper().startswith("SXM-") else f"SXM-{test_id}"
    
    def get_standardized_filename(test_id: str, file_type: str, extension: str) -> str:
        """Fallback get_standardized_filename function."""
        test_id = normalize_test_id(test_id)
        return f"{test_id}_{file_type}.{extension}"
    
    # Define a dummy OutputType enum
    class OutputType:
        PRIMARY_REPORT = "primary"
    
    def get_output_path(base_dir: str, test_id: str, filename: str, output_type=None) -> str:
        """Fallback get_output_path function."""
        return os.path.join(base_dir, filename)

class ExcelReportGenerator(ReportGenerator):
    """Generator for Excel reports."""
    
    def generate(self, data: ReportData) -> str:
        """
        Generate an Excel report.
        
        Args:
            data: Report data
            
        Returns:
            Path to the generated report
        """
        # Generate filename and path using utilities when available
        filename = get_standardized_filename(self.config.test_id, "log_analysis", "xlsx")
        excel_path = get_output_path(
            self.config.output_dir,
            self.config.test_id,
            filename,
            OutputType.PRIMARY_REPORT
        )
        
        # Check if Excel file is already open
        excel_file_available = True
        try:
            if os.path.exists(excel_path):
                with open(excel_path, 'a+b') as test_file:
                    pass
        except PermissionError:
            logging.warning(f"Excel file {excel_path} is open in another program. Will skip Excel output.")
            excel_file_available = False
        
        if not excel_file_available:
            return ""
        
        # Generate Excel report
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Create and write each sheet
            self._create_summary_sheet(writer, data.summary)
            
            if data.ymir_flag:
                self._create_scenario_sheet(writer, data.background_text, data.scenario_text)
                
            self._create_technical_summary_sheet(writer, data.errors)
            
            if data.component_analysis:
                self._create_component_analysis_sheet(writer, data.component_analysis)
                
            if data.ocr_data:
                self._create_images_sheet(writer, data.ocr_data)
                
            self._create_grouped_issues_sheet(writer, data.clusters)
            self._create_cluster_summary_sheet(writer, data.clusters)
                
        logging.info(f"Excel report written to: {excel_path}")
        return excel_path
    
    def _create_summary_sheet(self, writer, summary: str):
        """Create the Summary sheet."""
        summary_df = pd.DataFrame({
            'Test ID': [self.config.test_id],
            'AI-Generated Analysis': [summary]
        })
        # Sanitize the dataframe before writing to Excel
        summary_df = self._sanitize_dataframe(summary_df)
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format the sheet
        summary_sheet = writer.sheets['Summary']
        summary_sheet.row_dimensions[2].height = 400
        summary_sheet.column_dimensions['A'].width = 15
        summary_sheet.column_dimensions['B'].width = 85
        
        for col in range(1, 3):
            cell = summary_sheet.cell(row=1, column=col)
            self._format_cell(cell, wrap_text=False, bold=True, fill_color='D9D9D9')
            cell = summary_sheet.cell(row=2, column=col)
            self._format_cell(cell, wrap_text=True)
    
    def _create_scenario_sheet(self, writer, background_text: str, scenario_text: str):
        """Create the Scenario sheet for Ymir tests."""
        scenario_df = pd.DataFrame({
            'Test ID': [self.config.test_id],
            'Background': [background_text],
            'Test Scenario': [scenario_text]
        })
        # Sanitize the dataframe
        scenario_df = self._sanitize_dataframe(scenario_df)
        scenario_df.to_excel(writer, sheet_name='Scenario', index=False)
        
        # Format the sheet
        scenario_sheet = writer.sheets['Scenario']
        scenario_sheet.column_dimensions['A'].width = 15
        scenario_sheet.column_dimensions['B'].width = 40
        scenario_sheet.column_dimensions['C'].width = 60
        
        for col in range(1, 4):
            cell = scenario_sheet.cell(row=1, column=col)
            self._format_cell(cell, wrap_text=False, bold=True, fill_color='D9D9D9')
            cell = scenario_sheet.cell(row=2, column=col)
            self._format_cell(cell, wrap_text=True)
            
        scenario_sheet.row_dimensions[2].height = 200
    
    def _create_technical_summary_sheet(self, writer, errors: List[Dict]):
        """Create the Technical Summary sheet."""
        tech_errors = []
        for err in errors[:20]:
            if isinstance(err, dict):
                txt = err.get('text', '')
                if isinstance(txt, str):
                    tech_errors.append(txt)
                else:
                    tech_errors.append(str(txt))
        
        technical_summary = "\n\n".join(tech_errors)
        tech_summary_df = pd.DataFrame({
            'Test ID': [self.config.test_id],
            'Technical Summary': [technical_summary]
        })
        
        # Sanitize the dataframe
        tech_summary_df = self._sanitize_dataframe(tech_summary_df)
        tech_summary_df.to_excel(writer, sheet_name='Technical Summary', index=False)
        
        # Format the sheet
        tech_sheet = writer.sheets['Technical Summary']
        tech_sheet.row_dimensions[2].height = 300
        tech_sheet.column_dimensions['A'].width = 15
        tech_sheet.column_dimensions['B'].width = 85
        
        for col in range(1, 3):
            cell = tech_sheet.cell(row=1, column=col)
            self._format_cell(cell, wrap_text=False, bold=True, fill_color='D9D9D9')
            cell = tech_sheet.cell(row=2, column=col)
            self._format_cell(cell, wrap_text=True)
    
    def _create_component_analysis_sheet(self, writer, component_analysis: Dict[str, Any]):
        """Create the Component Analysis sheet."""
        # Extract data from component analysis
        component_distribution = component_analysis.get("component_error_counts", {})
        
        # Get component analysis files 
        component_files = component_analysis.get("analysis_files", {})
        file_names = {k: os.path.basename(v) for k, v in component_files.items() if v}
        
        # Create data for component analysis sheet
        component_data = [{
            'Test ID': self.config.test_id,
            'Total Components': len(component_distribution) if component_distribution else 1,
            'Components with Errors': len([comp for comp in component_distribution.keys() if comp != 'unknown']) if component_distribution else 1,
            'Identified Root Cause': self.config.primary_issue_component.upper(),
            'Error Propagation Diagram': file_names.get("error_propagation", "Available"),
            'Component Heatmap': file_names.get("error_heatmap", "Available"),
            'Component Report': file_names.get("component_report", "Available")
        }]
        
        component_df = pd.DataFrame(component_data)
        component_df = self._sanitize_dataframe(component_df)
        component_df.to_excel(writer, sheet_name='Component Analysis', index=False)
        
        # Format the sheet
        component_sheet = writer.sheets['Component Analysis']
        
        # Set column widths
        column_widths = {'A': 15, 'B': 18, 'C': 20, 'D': 25, 'E': 30, 'F': 30, 'G': 30}
        for col_letter, width in column_widths.items():
            component_sheet.column_dimensions[col_letter].width = width
        
        # Format header and data
        for col in range(1, 8):
            cell = component_sheet.cell(row=1, column=col)
            self._format_cell(cell, wrap_text=False, bold=True, fill_color='D9D9D9')
            if len(component_data) > 0:
                cell = component_sheet.cell(row=2, column=col)
                self._format_cell(cell)
                
        # Force the root cause cell value to primary component
        cell = component_sheet.cell(row=2, column=4)
        cell.value = self.config.primary_issue_component.upper()
        # Add visual emphasis to show it's the primary component
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    def _create_images_sheet(self, writer, ocr_data: List[Dict]):
        """Create the Images extraction sheet."""
        images_df = pd.DataFrame(ocr_data)
        images_df.rename(columns={'file': 'File Name', 'text': 'Extracted Text'}, inplace=True)
        
        # Sanitize the dataframe
        images_df = self._sanitize_dataframe(images_df)
        images_df.to_excel(writer, sheet_name='Images extraction', index=False)
        
        # Format the sheet
        images_sheet = writer.sheets['Images extraction']
        images_sheet.column_dimensions['A'].width = 30
        images_sheet.column_dimensions['B'].width = 100
        
        for col in range(1, 3):
            cell = images_sheet.cell(row=1, column=col)
            self._format_cell(cell, wrap_text=False, bold=True, fill_color='D9D9D9')
        
        for row_idx in range(2, len(ocr_data) + 2):
            images_sheet.row_dimensions[row_idx].height = 80
            for col_idx in range(1, 3):
                cell = images_sheet.cell(row=row_idx, column=col_idx)
                self._format_cell(cell)
    
    def _create_grouped_issues_sheet(self, writer, clusters: Dict[int, List[Dict]]):
        """Create the Grouped Issues sheet."""
        grouped_data = []
        cluster_colors = ['E6F0FF', 'F0E6FF', 'E6FFE6', 'FFE6E6', 'FFFDE6', 'E6FFFF', 'FFE6F0', 'E6F5FF']
        
        for cluster_id, errs in clusters.items():
            # Create a list for sorting using uniform datetime objects
            sorting_data = []
            for i, err in enumerate(errs):
                # Skip non-dictionary errors
                if not isinstance(err, dict):
                    continue
                    
                # Convert timestamp to datetime for sorting, if possible
                ts = err.get('timestamp', "No timestamp")
                dt_ts = ensure_datetime(ts) if ts != "No timestamp" else None
                sorting_data.append((i, err, dt_ts))
            
            # Sort by timestamp then by line number for stable ordering
            # First, sort entries with timestamps
            with_ts = sorted([x for x in sorting_data if x[2] is not None], key=lambda x: x[2])
            # Then, append entries without timestamps, sorted by line number
            without_ts = sorted([x for x in sorting_data if x[2] is None], 
                               key=lambda x: x[1].get('line_num', 0))
            
            # Combine sorted entries
            for i, err, _ in with_ts + without_ts:
                sequence_id = f"{cluster_id}.{i+1}"
                
                timestamp_val = err.get('timestamp')
                timestamp_info = ""
                if timestamp_val is not None:
                    if isinstance(timestamp_val, datetime):
                        timestamp_info = f"[{timestamp_val.isoformat()}] "
                    elif timestamp_val != "No timestamp":
                        timestamp_info = f"[{timestamp_val}] "
                
                # Always retrieve and use component information
                component_info = str(err.get('component', 'unknown'))
                
                # Convert error text to string if it's not
                error_text = err.get('text', '')
                if not isinstance(error_text, str):
                    error_text = str(error_text)
                
                grouped_data.append({
                    'Test ID': self.config.test_id,
                    'Cluster ID': cluster_id,
                    'Sequence': sequence_id,
                    'Timestamp': timestamp_val if isinstance(timestamp_val, str) else (
                        timestamp_val.isoformat() if isinstance(timestamp_val, datetime) else "No timestamp"
                    ),
                    'Severity': err.get('severity', 'Low'),
                    'Component': component_info,
                    'Log File': err.get('file', 'Unknown'),
                    'Line': err.get('line_num', 0),
                    'Error Message': timestamp_info + error_text
                })
        
        if not grouped_data:
            return
            
        grouped_errors_df = pd.DataFrame(grouped_data)
        # Sanitize the dataframe
        grouped_errors_df = self._sanitize_dataframe(grouped_errors_df)
        grouped_errors_df.to_excel(writer, sheet_name='Grouped Issues', index=False)
        grouped_sheet = writer.sheets['Grouped Issues']
        
        column_widths = {'A': 15, 'B': 10, 'C': 10, 'D': 18, 'E': 12, 'F': 15, 'G': 25, 'H': 10, 'I': 60}
        for col_letter, width in column_widths.items():
            grouped_sheet.column_dimensions[col_letter].width = width
        
        # Format headers
        for col in range(1, 10):
            cell = grouped_sheet.cell(row=1, column=col)
            self._format_cell(cell, wrap_text=False, bold=True, fill_color='D9D9D9')
        
        # Format data rows with cluster-based coloring - with additional safety
        prev_cluster = None
        for row_idx, row in enumerate(grouped_data, start=2):
            grouped_sheet.row_dimensions[row_idx].height = 60
            current_cluster = row['Cluster ID']
            if current_cluster != prev_cluster:
                cluster_color = cluster_colors[current_cluster % len(cluster_colors)]
                prev_cluster = current_cluster
            
            for col_idx, value in enumerate(row.values(), start=1):
                cell = grouped_sheet.cell(row=row_idx, column=col_idx)
                # Additional safety check for each value
                if isinstance(value, str):
                    cell.value = sanitize_text(value)
                else:
                    cell.value = str(value)
                
                # Consistent highlighting for primary issue component
                if col_idx == 6 and value == self.config.primary_issue_component:  # Component column with primary issue
                    self._format_cell(cell, fill_color='FFCCCC', bold=True)
                elif col_idx != 5:  # Not severity column
                    self._format_cell(cell, fill_color=cluster_color)
                else:  # Severity column
                    if value == 'High':
                        self._format_cell(cell, fill_color='FFB6B6')
                    elif value == 'Medium':
                        self._format_cell(cell, fill_color='FFE4B5')
                    else:
                        self._format_cell(cell, fill_color=cluster_color)
    
    def _create_cluster_summary_sheet(self, writer, clusters: Dict[int, List[Dict]]):
        """Create the Cluster Summary sheet."""
        cluster_summaries = []
        cluster_colors = ['E6F0FF', 'F0E6FF', 'E6FFE6', 'FFE6E6', 'FFFDE6', 'E6FFFF', 'FFE6F0', 'E6F5FF']
        
        for cluster_id, errs in clusters.items():
            # Skip clusters with no valid errors
            if not any(isinstance(e, dict) for e in errs):
                continue
                
            high_count = sum(1 for e in errs if isinstance(e, dict) and e.get('severity') == 'High')
            medium_count = sum(1 for e in errs if isinstance(e, dict) and e.get('severity') == 'Medium')
            low_count = sum(1 for e in errs if isinstance(e, dict) and e.get('severity') == 'Low')
            
            # Get representative error
            rep_error = None
            for e in errs:
                if isinstance(e, dict):
                    rep_error = e
                    break
            
            if not rep_error:
                continue
                
            error_text = rep_error.get('text', 'No error text')
            if not isinstance(error_text, str):
                error_text = str(error_text)
            error_sample = error_text[:100] + "..." if len(error_text) > 100 else error_text
            
            # Always retrieve and use component information consistently
            component_counts = {}
            for err in errs:
                if isinstance(err, dict):
                    component = err.get('component', 'unknown')
                    component_counts[component] = component_counts.get(component, 0) + 1
            
            primary_cluster_component = "unknown"
            if component_counts:
                primary_cluster_component = max(component_counts.items(), key=lambda x: x[1])[0]
            
            # Convert timestamps to datetime objects for min/max operations
            timestamps = []
            for e in errs:
                if not isinstance(e, dict):
                    continue
                ts = e.get('timestamp')
                if ts and ts != "No timestamp":
                    if isinstance(ts, datetime):
                        timestamps.append(ts)
                    else:
                        dt_ts = ensure_datetime(ts)
                        if dt_ts:
                            timestamps.append(dt_ts)
            
            timespan = "Unknown"
            if timestamps:
                min_ts = min(timestamps)
                max_ts = max(timestamps)
                timespan = f"{min_ts.isoformat()} to {max_ts.isoformat()}"
            
            cluster_summaries.append({
                'Cluster ID': cluster_id,
                'Error Count': len(errs),
                'High Severity': high_count,
                'Medium Severity': medium_count,
                'Low Severity': low_count,
                'Primary Component': primary_cluster_component,
                'Time Span': timespan,
                'Representative Error': error_sample
            })
        
        if not cluster_summaries:
            return
            
        # Sort clusters by importance
        cluster_summaries = sorted(
            cluster_summaries,
            key=lambda x: (x['High Severity'], x['Medium Severity'], x['Error Count']),
            reverse=True
        )
        
        cluster_summary_df = pd.DataFrame(cluster_summaries)
        # Sanitize the dataframe
        cluster_summary_df = self._sanitize_dataframe(cluster_summary_df)
        cluster_summary_df.to_excel(writer, sheet_name='Cluster Summary', index=False)
        summary_sheet_cs = writer.sheets['Cluster Summary']
        
        column_widths = {'A': 10, 'B': 12, 'C': 15, 'D': 15, 'E': 15, 'F': 15, 'G': 25, 'H': 60}
        for col_letter, width in column_widths.items():
            summary_sheet_cs.column_dimensions[col_letter].width = width
        
        # Format headers
        for col in range(1, 9):
            cell = summary_sheet_cs.cell(row=1, column=col)
            self._format_cell(cell, wrap_text=False, bold=True, fill_color='D9D9D9')
        
        # Format data rows with cluster-based coloring
        for row_idx, row in enumerate(cluster_summaries, start=2):
            summary_sheet_cs.row_dimensions[row_idx].height = 60
            cluster_color = cluster_colors[row['Cluster ID'] % len(cluster_colors)]
            
            for col_idx, value in enumerate(row.values(), start=1):
                cell = summary_sheet_cs.cell(row=row_idx, column=col_idx)
                # Additional safety check
                if isinstance(value, str):
                    cell.value = sanitize_text(value)
                else:
                    cell.value = str(value)
                    
                # Consistent highlighting of primary issue component
                if col_idx == 6 and value == self.config.primary_issue_component:  # Primary Component column
                    self._format_cell(cell, fill_color='FFCCCC', bold=True)
                else:
                    self._format_cell(cell, fill_color=cluster_color)
    
    def _sanitize_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sanitize a DataFrame to remove characters that are illegal in Excel worksheets."""
        # Create a new DataFrame to avoid modifying the original
        sanitized_df = pd.DataFrame()
        for column in df.columns:
            sanitized_df[column] = df[column].map(sanitize_text)
        return sanitized_df
    
    def _format_cell(self, cell, wrap_text=True, alignment='left', bold=False, fill_color=None, border=True):
        """Apply consistent formatting to Excel cells."""
        cell.alignment = Alignment(wrap_text=wrap_text, horizontal=alignment, vertical='top')
        if bold:
            cell.font = Font(bold=True)
        if fill_color:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        if border:
            side = Side(style='thin', color='000000')
            cell.border = Border(top=side, left=side, right=side, bottom=side)